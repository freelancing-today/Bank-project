from flask import Flask, request, redirect, render_template, send_file
from openpyxl import load_workbook
import openpyxl
from word2number import w2n
from decimal import Decimal, ROUND_DOWN
import re

app = Flask(__name__)
EXCEL_FILE = "output.xlsx"
# --- Formatting Utilities ---
def format_reference_number(reference_number):
    reference_number = reference_number.replace('\n', '').replace('\r', '')
    parts = reference_number.split()
    return '   '.join(parts)

def format_gaurantor_number(gaurantor_number):
    gaurantor_number = gaurantor_number.replace('\n', '').replace('\r', '')
    parts = gaurantor_number.split()
    return '   '.join(parts)

def format_name(name):
    parts = name.strip().upper().split()
    return "  ".join(parts)

def format_ref(ref):
    return ref.replace(" ", "   ").replace("\n", "")

def format_city_state(cs):
    parts = cs.strip().upper().split(',')
    if len(parts) == 2:
        return f"{parts[0].strip()} , {parts[1].strip()}"
    return cs.upper()

def format_currency(val):
    s = f"{val:,.2f}"
    return "$  " + s.replace(",", " , ")

def format_percent(val):
    return f"{val:.2f} %"

# --- Alphanumeric to Numeric Conversion ---


# --- Verified Manual Calculation ---
def calculate_all(purchase_value, pur_red_pct, down_pct, loan_period, interest_pct, mon_prin_red_pct, total_int_red_pct):
    reduced = purchase_value * (pur_red_pct / 100)
    reduced = float(Decimal(reduced).quantize(Decimal("0.01"), rounding=ROUND_DOWN))
    final_purchase = purchase_value - reduced
    final_purchase = float(Decimal(final_purchase).quantize(Decimal("0.01"), rounding=ROUND_DOWN))

    down_payment = final_purchase * (down_pct / 100)
    down_payment = float(Decimal(down_payment).quantize(Decimal("0.01"), rounding=ROUND_DOWN))
    loan_amt = final_purchase - down_payment
    loan_amt = float(Decimal(loan_amt).quantize(Decimal("0.01"), rounding=ROUND_DOWN))

    annual = loan_amt / loan_period
    monthly = annual / 12
    principal = monthly * (mon_prin_red_pct / 100)
    principal = float(Decimal(principal).quantize(Decimal("0.01"), rounding=ROUND_DOWN))

    annual_interest = loan_amt * (interest_pct / 100)
    total_interest = annual_interest * loan_period
    final_interest = total_interest * (total_int_red_pct / 100)
    final_interest = float(Decimal(final_interest).quantize(Decimal("0.01"), rounding=ROUND_DOWN))

    insurance = loan_amt * 0.0032
    insurance_monthly = insurance / 12
    insurance_monthly = float(Decimal(insurance_monthly).quantize(Decimal("0.01"), rounding=ROUND_DOWN))

    loan_pct = 100 - down_pct
    if loan_pct <= 80 or loan_period > 20:
        pmi = "NA"
    else:
        pmi_rate = 0.19 if 80.01 <= loan_pct <= 85 else 0.23 if 85.01 <= loan_pct <= 90 else 0.26 if 90.01 <= loan_pct <= 95 else "NA"
        pmi = format_currency(loan_amt * (pmi_rate / 100)) if isinstance(pmi_rate, float) else "NA"

    return {
        "final_purchase": final_purchase,
        "down_payment": down_payment,
        "loan_amt": loan_amt,
        "principal": principal,
        "final_interest": final_interest,
        "insurance_monthly": insurance_monthly,
        "pmi": pmi
    } 
# --- Flask Routes ---
@app.route("/")
def home():
    return render_template("form.html")  # Use render_template to load the form.html file

@app.route("/submit", methods=["GET", "POST"])
def submit():
    if request.method == "GET":
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            if ws.max_row <= 1:  # Check if there are no entries (only header row exists)
                return "NO LOAN ENTRIES FOUND"
        except FileNotFoundError:
            return "NO LOAN ENTRIES FOUND"
        return render_template("submit.html")  # Render the page for GET requests

    # Handle POST request
    data = request.form

    customer_ref = format_ref(data['customerRef'])
    customer_name = format_name(data['customerName'])
    city_state = format_city_state(data['cityState'])
    guarantor_name = format_name(data['guarantorName'])
    guarantor_ref = format_ref(data['guarantorRef'])

    purchase_value_words = request.form['purchaseValue']
    purchase_reduction = float(request.form['purchaseReduction'])

    # Clean the input (e.g., "$ ... dollars and ... cents")
    cleaned = purchase_value_words.lower().replace("$", "").replace("dollars", "").replace("and", "").replace("cents", "").strip()
    
    # Split integer and decimal parts
    if 'point' in cleaned:
        words_parts = cleaned.split("point")
        int_part = w2n.word_to_num(words_parts[0].strip())
        dec_part = float("0." + ''.join(words_parts[1].split()))
    elif 'dot' in cleaned:
        words_parts = cleaned.split("dot")
        int_part = w2n.word_to_num(words_parts[0].strip())
        dec_part = float("0." + ''.join(words_parts[1].split()))
    else:
        words_parts = cleaned.split()
        if 'cent' in words_parts:
            idx = words_parts.index('cent')
            int_part = w2n.word_to_num(" ".join(words_parts[:idx]))
            dec_part = float("0." + ''.join(words_parts[idx+1:]))
        else:
            int_part = w2n.word_to_num(cleaned)
            dec_part = 0.0

    numeric_value = int_part + dec_part
    purchase_value = numeric_value
    # Apply reduction
    reduction_amount = (numeric_value * purchase_reduction) / 100
    reduction_amount = float(str(reduction_amount)[:str(reduction_amount).find('.') + 3])  # precise 2 decimals

    final_value = numeric_value - reduction_amount
    final_value = float(str(final_value)[:str(final_value).find('.') + 3])

    pur_red = float(data["purchaseReduction"])  # Purchase reduction percentage
    down_pct = float(data["downPayment"])  # Down payment percentage
    loan_years = int(data["loanPeriod"])  # Loan period (years)
    interest = float(data["annualInterest"])  # Annual interest percentage
    mon_prin_red = float(data["monthlyPrincipalReduction"])  # Monthly principal reduction percentage
    total_int_red = float(data["totalInterestReduction"])  # Total interest reduction percentage

    # Perform all calculations
    calcs = calculate_all(purchase_value, pur_red, down_pct, loan_years, interest, mon_prin_red, total_int_red)

    # Load the Excel workbook and prepare for saving the results
    try:
        wb = load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        # Create a new Excel file if it doesn't exist
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Entrier Name", "Customer Ref", "Customer Name", "City, State", "Purchase Value and Down Payment", "Loan Period and Annual Interest", "Guarantor Name", "Guarantor Ref", "Loan Amount and Principal", "Total Interest and Property Tax", "Property Insurance and PMI"])
        wb.save(EXCEL_FILE)
    ws = wb.active
    row = ws.max_row + 1
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=4).value = final_value
    wb.save(EXCEL_FILE)

    # Convert down payment to integer before saving to Excel
    down_payment_amount = int(data['downPayment'])

    # Write results to the Excel file
    ws.cell(row=row, column=1, value=customer_ref)
    ws.cell(row=row, column=2, value=customer_name)
    ws.cell(row=row, column=3, value=city_state)
    ws.cell(row=row, column=4, value=f"{format_currency(calcs['final_purchase'])} AND {down_payment_amount}")
    ws.cell(row=row, column=5, value=f"{loan_years} YEARS AND {format_percent(interest)}")
    ws.cell(row=row, column=6, value=guarantor_name)
    ws.cell(row=row, column=7, value=guarantor_ref)
    ws.cell(row=row, column=8, value=f"{format_currency(calcs['loan_amt'])} AND {format_currency(calcs['principal'])}")
    ws.cell(row=row, column=9, value=f"{format_currency(calcs['final_interest'])} AND")
    ws.cell(row=row, column=10, value=f"{format_currency(calcs['insurance_monthly'])} AND {calcs['pmi']}")

    wb.save(EXCEL_FILE)
    return redirect("/")

@app.route("/download", methods=["GET"])
def download():
    try:
        return send_file("output.xlsx", as_attachment=True)
    except FileNotFoundError:
        return "The requested file was not found on the server.", 404

if __name__ == "__main__":
    app.run(debug=True)
